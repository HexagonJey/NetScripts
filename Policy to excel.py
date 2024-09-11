import os

import openpyxl
from ipaddress import IPv4Network

filename = 'C:\\Users\\"UserID"\\Downloads\\Scripts\\"DeviceName"-Config.txt'

storepath = os.getcwd() 

#filename = storepath + '\\' + 'config_file_policy.txt'

junos_app = ['junos-ftp', 'junos-tftp', 'junos-rtsp', 'junos-netbios-session', 'junos-smb-session', 'junos-ssh', 'junos-telnet', 'junos-smtp', 'junos-tacacs', 'junos-tacacs-ds', 'junos-dhcp-client', 'junos-dhcp-server', 'junos-bootpc', 'junos-bootps', 'junos-finger', 'junos-http', 'junos-https', 'junos-pop3', 'junos-ident', 'junos-nntp', 'junos-ntp', 'junos-imap', 'junos-imaps', 'junos-bgp', 'junos-ldap', 'junos-snpp', 'junos-biff', 'junos-who', 'junos-syslog', 'junos-printer', 'junos-rip', 'junos-radius', 'junos-radacct', 'junos-nfsd-tcp', 'junos-nfsd-udp', 'junos-cvspserver', 'junos-ldp-tcp', 'junos-ldp-udp', 'junos-xnm-ssl', 'junos-xnm-clear-text', 'junos-ike', 'any', 'junos-aol', 'junos-chargen', 'junos-dhcp-relay', 'junos-discard', 'junos-dns-udp', 'junos-dns-tcp', 'junos-echo', 'junos-gopher', 'junos-gnutella', 'junos-gre', 'junos-gprs-gtp-c-tcp', 'junos-gprs-gtp-c-udp', 'junos-gprs-gtp-c', 'junos-gprs-gtp-c', 'junos-gprs-gtp-u-tcp', 'junos-gprs-gtp-u-udp', 'junos-gprs-gtp-u', 'junos-gprs-gtp-u', 'junos-gprs-gtp-v0-tcp', 'junos-gprs-gtp-v0-udp', 'junos-gprs-gtp-v0', 'junos-gprs-gtp-v0', 'junos-gprs-sctp', 'junos-http-ext', 'junos-icmp-all', 'junos-icmp-ping', 'junos-internet-locator-service', 'junos-ike-nat', 'junos-irc', 'junos-l2tp', 'junos-lpr', 'junos-mail', 'junos-h323', 'junos-h323', 'junos-h323', 'junos-h323', 'junos-h323', 'junos-h323', 'junos-mgcp-ua', 'junos-mgcp-ca', 'junos-msn', 'junos-ms-rpc-tcp', 'junos-ms-rpc-udp', 'junos-ms-rpc-epm', 'junos-ms-rpc-msexchange-directory-rfr', 'junos-ms-rpc-msexchange-info-store', 'junos-ms-rpc-msexchange-directory-nsp', 'junos-ms-rpc-wmic-admin', 'junos-ms-rpc-wmic-webm-level1login', 'junos-ms-rpc-wmic-admin2', 'junos-ms-rpc-wmic-mgmt', 'junos-ms-rpc-iis-com-1', 'junos-ms-rpc-iis-com-adminbase', 'junos-ms-rpc-uuid-any-tcp', 'junos-ms-rpc-uuid-any-udp', 'junos-ms-sql', 'junos-nbname', 'junos-nbds', 'junos-nfs', 'junos-ns-global', 'junos-ns-global-pro', 'junos-nsm', 'junos-ospf', 'junos-pc-anywhere', 'junos-ping', 'junos-pingv6', 'junos-icmp6-dst-unreach-addr', 'junos-icmp6-dst-unreach-admin', 'junos-icmp6-dst-unreach-beyond', 'junos-icmp6-dst-unreach-port', 'junos-icmp6-dst-unreach-route', 'junos-icmp6-echo-reply', 'junos-icmp6-echo-request', 'junos-icmp6-packet-to-big', 'junos-icmp6-param-prob-header', 'junos-icmp6-param-prob-nexthdr', 'junos-icmp6-param-prob-option', 'junos-icmp6-time-exceed-reassembly', 'junos-icmp6-time-exceed-transit', 'junos-icmp6-all', 'junos-pptp', 'junos-realaudio', 'junos-sccp', 'junos-sctp-any', 'junos-sip', 'junos-sip', 'junos-rsh', 'junos-smb', 'junos-smb', 'junos-sql-monitor', 'junos-sqlnet-v1', 'junos-sqlnet-v2', 'junos-sun-rpc-tcp', 'junos-sun-rpc-udp', 'junos-sun-rpc-portmap-tcp', 'junos-sun-rpc-portmap-udp', 'junos-sun-rpc-nfs-tcp', 'junos-sun-rpc-nfs-udp', 'junos-sun-rpc-mountd-tcp', 'junos-sun-rpc-mountd-udp', 'junos-sun-rpc-ypbind-tcp', 'junos-sun-rpc-ypbind-udp', 'junos-sun-rpc-status-tcp', 'junos-sun-rpc-status-udp', 'junos-talk', 'junos-talk', 'junos-ntalk', 'junos-ntalk', 'junos-tcp-any', 'junos-udp-any', 'junos-uucp', 'junos-vdo-live', 'junos-vnc', 'junos-wais', 'junos-whois', 'junos-winframe', 'junos-x-windows', 'junos-ymsg', 'junos-wxcontrol', 'junos-snmp-agentx', 'junos-stun', 'junos-stun', 'junos-persistent-nat', 'junos-r2cp']
default_app = ['ftp', 'tftp', 'rtsp', 'netbios', 'smb', 'ssh', 'telnet', 'smtp', 'tacacs', 'tacacs', 'dhcp', 'dhcp', 'bootpc', 'bootps', 'finger', 'http', 'https', 'pop3', 'ident', 'nntp', 'ntp', 'imap', 'imaps', 'bgp', 'ldap', 'snpp', 'biff', 'who', 'syslog', 'printer', 'rip', 'radius', 'radacct', 'nfsd', 'nfsd', 'cvspserver', 'ldp', 'ldp', 'xnm', 'xnm', 'ike', 'any', 'aol', 'chargen', 'dhcp', 'discard', 'dns', 'dns', 'echo', 'gopher', 'gnutella', 'gre', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'gprs', 'http', 'icmp', 'icmp', 'internet', 'ike', 'irc', 'l2tp', 'lpr', 'mail', 'h323', 'h323', 'h323', 'h323', 'h323', 'h323', 'mgcp', 'mgcp', 'msn', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'ms', 'nbname', 'nbds', 'nfs', 'ns', 'ns', 'nsm', 'ospf', 'pc', 'ping', 'pingv6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'icmp6', 'pptp', 'realaudio', 'sccp', 'sctp', 'sip', 'sip', 'rsh', 'smb', 'smb', 'sql', 'sqlnet', 'sqlnet', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'sun', 'talk', 'talk', 'ntalk', 'ntalk', 'tcp', 'udp', 'uucp', 'vdo', 'vnc', 'wais', 'whois', 'winframe', 'x', 'ymsg', 'wxcontrol', 'snmp', 'stun', 'stun', 'persistent', 'r2cp']
default_protocol = ['tcp', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'udp', 'udp', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'udp', 'udp', 'tcp', 'udp', 'udp', 'udp', 'tcp', 'udp', 'tcp', 'tcp', 'udp', 'tcp', 'tcp', 'udp', '0', '6', 'udp', 'udp', 'udp', 'udp', 'tcp', 'udp', 'tcp', 'udp', '47', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', '132', 'tcp', 'icmp', 'icmp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'tcp', 'tcp', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'udp', 'tcp', 'tcp', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'tcp', 'udp', 'udp', 'udp', 'tcp', 'tcp', 'udp', '89', 'udp', '1', '58', '58', '58', '58', '58', '58', '58', '58', '58', '58', '58', '58', '58', '58', '58', 'tcp', 'tcp', 'tcp', '132', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'tcp', 'tcp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'tcp', 'udp', 'udp', 'tcp', 'udp', 'tcp', 'tcp', 'udp', 'udp', 'udp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'tcp', 'udp', 'tcp', '255', 'udp']
default_port = ['21', '69', '554', '139', '445', '22', '23', '25', '49', '65', '68', '67', '68', '67', '79', '80', '443', '110', '113', '119', '123', '143', '993', '179', '389', '444', '512', '513', '514', '515', '520', '1812', '1813', '2049', '2049', '2401', '646', '646', '3220', '3221', '500', ' ', '5190-5193', '19', '67', '9', '53', '53', '7', '70', '6346-6347', ' ', '2123', '2123', '2123', '2123', '2152', '2152', '2152', '2152', '3386', '3386', '3386', '3386', '0', '7001', ' ', ' ', '389', '4500', '6660-6669', '1701', '515', '25', '1720', '1719', '1503', '389', '522', '1731', '2427', '2727', '1863', '135', '135', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', '1433', '137', '138', '111', '15397', '15397', '69', ' ', '5632', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', '1723', '554', '2000', ' ', '5060', '5060', '514', '139', '445', '1434', '1525', '1521', '111', '111', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', '517', '517', '518', '518', ' ', ' ', '540', '7000-7010', '5800', '210', '43', '1494', '6000-6063', '5050', '3578', '705', '3478-3479', '3478-3479', '65535', '28672']

Add_obj = []
Ip_add = []
IP_add_x_mask = []
IP_add_mask_o = []
Ip_add_obj_in_group = []
Add_obj_grp = []
App_nm = []
App_prt = []
App_dest_prt = []
App_grp_nm = []
App_grp_mm = []

with open (filename) as device:
    config = device.read().splitlines()
    print (config)
    for idx,lines in enumerate(config):
        if 'address-book address ' in lines:
            linessplit = lines.split()
            Add_obj.append(linessplit[9])
            Ip_add.append(linessplit[10])

        elif 'address-book address-set' in lines:
            linessplit = lines.split()
            Add_obj_grp.append(linessplit[9])
            Ip_add_obj_in_group.append(linessplit[11])
        
        elif 'applications application ' and 'protocol '  in lines:
            linessplit = lines.split()
            #App_nm.append(linessplit[5])
            #App_prt.append(linessplit[7])
            linessplit = lines.split()
            idex = linessplit.index('application')
            App_nm.append(linessplit[idex + 1])
            idex2 = linessplit.index('protocol')
            App_prt.append(linessplit[idex2 + 1])
        
        elif 'applications application ' and 'destination-port '  in lines:
            linessplit = lines.split()
            #App_dest_prt.append(linessplit[7])
            idex3 = linessplit.index('destination-port')
            App_dest_prt.append(linessplit[idex3 + 1])
        
        elif 'applications application-set ' in lines:
            linessplit = lines.split()
            #App_grp_nm.append(linessplit[5])
            #App_grp_mm.append(linessplit[7])
            idex4 = linessplit.index('application-set')
            App_grp_nm.append(linessplit[idex4 + 1])
            App_grp_mm.append(linessplit[idex4 + 3])
        
            
for idx,address in enumerate(Ip_add):
    IPaddr = str(IPv4Network(Ip_add[idx]).network_address)
    SubnetMask = str(IPv4Network(Ip_add[idx]).netmask)
    IP_add_x_mask.append(IPaddr)
    IP_add_mask_o.append(SubnetMask)

excel = openpyxl.Workbook()

excel_ws = excel.create_sheet(title = 'Policy', index=0)
excel_ws = excel.get_sheet_by_name('Policy')

IP_sheet_ws = excel.create_sheet(title = 'Obj_Ip_Subnet', index=1)
IP_sheet_ws = excel.get_sheet_by_name('Obj_Ip_Subnet')

IP_sheet_ws.append(['IP Object', 'IP', 'Subnet'])

for idx,obj in enumerate(Add_obj):
        IP_sheet_ws.append([Add_obj[idx], IP_add_x_mask[idx], IP_add_mask_o[idx]])
        
App_sheet_ws = excel.create_sheet(title = 'App_Protocol_Port', index=2)
App_sheet_ws = excel.get_sheet_by_name('App_Protocol_Port')

App_sheet_ws.append(['Appliation', 'Protocol', 'Port From', 'Port To'])

for idx,obj in enumerate(App_nm):
    inc_idx = App_dest_prt[idx].find('-')
    dst_post_split = App_dest_prt[idx].split('-')
    if inc_idx == -1:
        App_sheet_ws.append([App_nm[idx], App_prt[idx], App_dest_prt[idx]])
    else:
        App_sheet_ws.append([App_nm[idx], App_prt[idx], dst_post_split[0], dst_post_split[1]])
    
with open (filename, 'r') as device:
    config = device.read()
    if not config.startswith('deactivate'):   
        firewall_rules = config.split('set logical-systems')

firewall_rules.pop(0)

extracted_rules = []
dest_address_g = []
dest_address_a = []
dest_address_ip = []
dest_address_ip_m = []
dest_address_g_x = ''
dest_address_a_x = ''
dest_address_ip_x = ''
dest_address_ip_m_x = ''
source_address_g = []
source_address_a = []
source_address_ip = []
source_address_ip_m = []
source_address_g_x = ''
source_address_a_x = ''
source_address_ip_x = ''
source_address_ip_m_x = ''
application_g = []
application_g_x = ''
application_m = []
application_m_x = ''
application_p = []
application_p_x = ''
application_dp = []
application_dp_x = '' 
policy = ''
from_zone = ''
to_zone = ''
logical_system = ''

for rules in firewall_rules:
    rules_lines = rules.split('/n')

    if 'policy' in rules:
        #if policy != '':
            #print (extracted_rules)
        extracted_rules.append([logical_system, policy, from_zone, to_zone, source_address_g_x, source_address_a_x, source_address_ip_x, source_address_ip_m_x ,  dest_address_g_x, dest_address_a_x , dest_address_ip_x, dest_address_ip_m_x, application_g_x, application_m_x, application_p_x, application_dp_x])

        logical_system = rules_lines[0].split()[0]
        from_zone = rules_lines[0].split()[4]
        to_zone = rules_lines[0].split()[6]
        policy = rules_lines[0].split()[8]
        
        dest_address_g = []
        dest_address_a = []
        dest_address_ip = []
        dest_address_ip_m = []
        dest_address_g_x = ''
        dest_address_a_x = ''
        dest_address_ip_x = ''
        dest_address_ip_m_x = ''
        source_address_g = []
        source_address_a = []
        source_address_ip = []
        source_address_ip_m = []
        source_address_g_x = ''
        source_address_a_x = ''
        source_address_ip_x = ''
        source_address_ip_m_x = ''
        application_g = []
        application_g_x = ''
        application_m = []
        application_m_x = ''
        application_p = []
        application_p_x = ''
        application_dp = []
        application_dp_x = ''        

        if 'policy' and 'source-address ' in rules:
            if rules_lines[0].split()[11] in Add_obj_grp:
                source_address_g.append(rules_lines[0].split()[11])
                for idx in range(len(Add_obj_grp)):
                    if Add_obj_grp[idx] == rules_lines[0].split()[11]:
                        source_address_a.append(Ip_add_obj_in_group[idx])
                        #print (Ip_add_obj[idx])
                        idx2 = (Add_obj.index(Ip_add_obj_in_group[idx]))
                        source_address_ip.append(IP_add_x_mask[idx2])
                        source_address_ip_m.append(IP_add_mask_o[idx2])
                    
                    elif rules_lines[0].split()[11] in Add_obj:
                        source_address_a.append(rules_lines[0].split()[11])
                        idx = Add_obj.index(rules_lines[0].split()[11])
                        source_address_ip.append(IP_add_x_mask[idx])
                        source_address_ip_m.append(IP_add_mask_o[idx])
                                                                                                    
        elif 'policy' and 'destination-address ' in rules: 
            if rules_lines[0].split()[11] in Add_obj_grp:
                dest_address_g.append(rules_lines[0].split()[11])
                for idx in range(len(Add_obj_grp)):
                    if Add_obj_grp[idx] == rules_lines[0].split()[11]:
                        dest_address_a.append(Ip_add_obj_in_group[idx])
                        #print (Ip_add_obj[idx])
                        idx2 = (Add_obj.index(Ip_add_obj_in_group[idx]))
                        dest_address_ip.append(IP_add_x_mask[idx2])
                        dest_address_ip_m.append(IP_add_mask_o[idx2])
                    
                    elif rules_lines[0].split()[11] in Add_obj:
                        dest_address_a.append(rules_lines[0].split()[11])
                        idx = Add_obj.index(rules_lines[0].split()[11])
                        dest_address_ip.append(IP_add_x_mask[idx])
                        dest_address_ip_m.append(IP_add_mask_o[idx])
    
        elif 'policy' in rules and 'application' in rules:
            print (rules)
            if rules_lines[0].split()[11] in App_grp_nm:
                application_g.append(rules_lines[0].split()[11])
                for idx in range(len(App_grp_nm)):
                    if App_grp_nm[idx] == rules_lines[0].split()[11]:
                        application_m.append(App_grp_mm[idx])
                        idx2 = (App_nm.index(App_grp_mm[idx]))
                        application_p.append(App_prt[idx2])
                        application_dp.append(App_dest_prt[idx2])
                        
                    elif rules_lines[0].split()[11] in junos_app:
                        idx = (junos_app.index(rules_lines[0].split()[11]))
                        application_m.append(default_app[idx])
                        application_p.append(default_protocol[idx])
                        application_dp.append(default_port[idx])        
                        
                    elif rules_lines[0].split()[11] in App_grp_mm:
                        application_m.append(rules_lines[0].split()[11])
                        idx = (App_nm.index(rules_lines[0].split()[11]))
                        application_p.append(App_prt[idx])
                        application_dp.append(App_dest_prt[idx])

                
        source_address_g_x ='\n'.join(source_address_g)
        source_address_a_x = '\n'.join(source_address_a)
        source_address_ip_x = '\n'.join(source_address_ip)
        source_address_ip_m_x = '\n'.join(source_address_ip_m)
        dest_address_g_x ='\n'.join(dest_address_g)
        dest_address_a_x = '\n'.join(dest_address_a)
        dest_address_ip_x = '\n'.join(dest_address_ip)
        dest_address_ip_m_x = '\n'.join(dest_address_ip_m)    
        application_g_x ='\n'.join(application_g)
        application_m_x ='\n'.join(application_m)
        application_p_x ='\n'.join(application_p)
        application_dp_x ='\n'.join(application_dp)
    
    #extracted_rules.append([policy, from_zone, to_zone, source_address_g_x, source_address_a_x, source_address_ip_x, source_address_ip_m_x ,  dest_address_g_x, dest_address_a_x , dest_address_ip_x, dest_address_ip_m_x, application_g_x, application_m_x, application_p_x, application_dp_x])
    #print(extracted_rules)

extracted_rules.append([policy, from_zone, to_zone, source_address_g_x, source_address_a_x, source_address_ip_x, source_address_ip_m_x ,  dest_address_g_x, dest_address_a_x , dest_address_ip_x, dest_address_ip_m_x, application_g_x, application_m_x, application_p_x, application_dp_x])

#excel_ws = excel.create_sheet(title = 'Policy', index=0)
#excel_ws = excel.get_sheet_by_name('Policy')

excel_ws.append(['Logical System', 'Policy', 'From Zone', 'To Zone', 'Source Address Group',  'Source Address Object', 'Source Address IP Address', 'Source Address Subnet Mask' , 'Destination Address Group', 'Destination Address Object', 'Destination IP Address', 'Destination IP Subnet Mask'  , 'Application Group',  'Application Member', 'Protocol', 'Destination Port'])

for row in extracted_rules:
   excel_ws.append(row)

excel.save('C:\\Users\\"UserID"\\Downloads\\Scripts\\srx_config.xlsx')

#excel.save(storepath + '\\' + 'srx_config.xlsx')
